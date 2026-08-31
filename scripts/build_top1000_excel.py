#!/usr/bin/env python3
"""
Build spanish-top-1000-words.xlsx
- Top 1000 frequency surface forms (subtitle corpus / Wiktionary list)
- Lemma, type, gender, English meaning, pronunciation
- For verbs: present tense 1st/2nd/3rd person singular + plural
"""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

SCRIPTS = Path(__file__).resolve().parent
ROOT = SCRIPTS.parent
FREQ = ROOT / "data" / "_freq1000.tsv"
OUT = ROOT / "spanish-top-1000-words.xlsx"

# ---------------------------------------------------------------------------
# Meanings for the main Top_1000 sheet (can include extra context)
# ---------------------------------------------------------------------------
MEANINGS: dict[str, str] = {
    "que": "that / which / than",
    "de": "of / from",
    "no": "no / not",
    "a": "to / at",
    "la": "the (f.) / her / it",
    "el": "the (m.)",
    "ser": "to be (identity)",
    "y": "and",
    "en": "in / on / at",
    "lo": "it / the (neuter)",
    "un": "a / an (m.)",
    "por": "for / by / through",
    "qué": "what?",
    "me": "me / myself",
    "te": "you / yourself",
    "se": "himself / herself / themselves",
    "con": "with",
    "para": "for / in order to",
    "mi": "my",
    "estar": "to be (state/location)",
    "si": "if",
    "bien": "well / fine / good",
    "pero": "but",
    "yo": "I",
    "eso": "that (neuter / that thing)",
    "sí": "yes",
    "su": "his / her / your / their",
    "tu": "your",
    "aquí": "here",
    "del": "of the (de+el)",
    "al": "to the (a+el)",
    "como": "like / as",
    "le": "to him / her / you",
    "más": "more",
    "esto": "this (neuter / this thing)",
    "ya": "already / now",
    "todo": "all / everything",
    "esta": "this (f.)",
    "ir": "to go",
    "muy": "very",
    "haber": "to have (aux.) / there is",
    "ahora": "now",
    "algo": "something",
    "tener": "to have",
    "nos": "us",
    "tú": "you (informal)",
    "nada": "nothing",
    "cuando": "when",
    "este": "this (m.) / east",
    "saber": "to know (facts)",
    "así": "like this / so",
    "poder": "to be able to / can",
    "cómo": "how?",
    "querer": "to want / to love",
    "sólo": "only",
    "solo": "alone / only",
    "gracias": "thank you",
    "o": "or",
    "él": "he",
    "bueno": "good",
    "hacer": "to do / to make",
    "vez": "time (occasion)",
    "creer": "to believe",
    "ella": "she",
    "ese": "that (m.)",
    "esa": "that (f.)",
    "esos": "those (m.)",
    "esas": "those (f.)",
    "estos": "these (m.)",
    "estas": "these (f.)",
    "aquel": "that (m., farther away)",
    "aquella": "that (f., farther away)",
    "aquellos": "those (m., farther away)",
    "aquellas": "those (f., farther away)",
    "aquello": "that (neuter, farther away)",
    "hola": "hello",
    "porque": "because",
    "dios": "god",
    "quién": "who?",
    "nunca": "never",
    "dónde": "where?",
    "casa": "house / home",
    "casar": "to marry",
    "favor": "favor",
    "dos": "two",
    "tanto": "so much",
    "señor": "sir / Mr.",
    "tiempo": "time / weather",
    "verdad": "truth",
    "mejor": "better / best",
    "hombre": "man",
    "usted": "you (formal)",
    "mucho": "much / a lot",
    "entonces": "then / so",
    "sentir": "to feel / to regret",
    "sentar": "to seat",
    "ahí": "there",
    "ti": "you (after prep.)",
    "vida": "life",
    "ver": "to see",
    "alguien": "someone",
    "hasta": "until / even",
    "sin": "without",
    "mí": "me (after prep.)",
    "año": "year",
    "sobre": "on / about",
    "decir": "to say / to tell",
    "uno": "one",
    "siempre": "always",
    "oh": "oh",
    "cosa": "thing",
    "también": "also",
    "antes": "before",
    "ni": "nor / not even",
    "día": "day",
    "noche": "night",
    "nadie": "nobody",
    "otro": "other",
    "parecer": "to seem",
    "nosotros": "we",
    "poco": "little / few",
    "padre": "father",
    "trabajo": "work / job",
    "trabajar": "to work",
    "gente": "people",
    "mirar": "to look",
    "les": "to them / to you all",
    "donde": "where",
    "mismo": "same",
    "ellos": "they (m.)",
    "pasar": "to pass / to happen",
    "dinero": "money",
    "hijo": "son / child",
    "tal": "such",
    "hablar": "to speak",
    "seguro": "sure / safe",
    "claro": "clear / of course",
    "lugar": "place",
    "mundo": "world",
    "amigo": "friend (m.)",
    "esperar": "to wait / to hope",
    "después": "after",
    "momento": "moment",
    "desde": "from / since",
    "tipo": "type / guy",
    "mañana": "tomorrow / morning",
    "gran": "great / big",
    "grande": "big",
    "necesitar": "to need",
    "estado": "state / been",
    "acuerdo": "agreement",
    "papá": "dad",
    "mío": "mine",
    "gustar": "to like (please)",
    "nuestro": "our",
    "nuevo": "new",
    "nombre": "name",
    "tres": "three",
    "menos": "less",
    "deber": "should / to owe",
    "mal": "bad / badly",
    "conmigo": "with me",
    "madre": "mother",
    "hoy": "today",
    "quien": "who",
    "mamá": "mom",
    "luego": "then / later",
    "allí": "there",
    "hora": "hour",
    "mujer": "woman",
    "tarde": "afternoon / late",
    "oír": "to hear",
    "parte": "part",
    "problema": "problem",
    "mas": "but (literary)",
    "importar": "to matter",
    "contigo": "with you",
    "venir": "to come",
    "pensar": "to think",
    "llegar": "to arrive",
    "llevar": "to carry / to wear",
    "dejar": "to leave / to let",
    "quedar": "to remain / to meet",
    "llamar": "to call",
    "tomar": "to take / to drink",
    "conocer": "to know (people)",
    "vivir": "to live",
    "contar": "to count / to tell",
    "empezar": "to begin",
    "buscar": "to look for",
    "existir": "to exist",
    "entrar": "to enter",
    "escribir": "to write",
    "perder": "to lose",
    "producir": "to produce",
    "ocurrir": "to occur",
    "entender": "to understand",
    "pedir": "to ask for",
    "recibir": "to receive",
    "recordar": "to remember",
    "terminar": "to finish",
    "permitir": "to allow",
    "aparecer": "to appear",
    "conseguir": "to get / achieve",
    "comenzar": "to begin",
    "servir": "to serve",
    "sacar": "to take out",
    "mantener": "to maintain",
    "resultar": "to result / turn out",
    "leer": "to read",
    "caer": "to fall",
    "cambiar": "to change",
    "presentar": "to present",
    "crear": "to create",
    "abrir": "to open",
    "considerar": "to consider",
    "acabar": "to finish",
    "convertir": "to convert",
    "ganar": "to win / earn",
    "formar": "to form",
    "traer": "to bring",
    "partir": "to leave / divide",
    "morir": "to die",
    "aceptar": "to accept",
    "realizar": "to carry out",
    "suponer": "to suppose",
    "comprender": "to understand",
    "lograr": "to achieve",
    "explicar": "to explain",
    "preguntar": "to ask",
    "tocar": "to touch / play",
    "reconocer": "to recognize",
    "estudiar": "to study",
    "alcanzar": "to reach",
    "nacer": "to be born",
    "dirigir": "to direct",
    "correr": "to run",
    "utilizar": "to use",
    "pagar": "to pay",
    "ayudar": "to help",
    "jugar": "to play",
    "escuchar": "to listen",
    "cumplir": "to fulfill",
    "ofrecer": "to offer",
    "descubrir": "to discover",
    "levantar": "to raise",
    "intentar": "to attempt",
    "usar": "to use",
    "decidir": "to decide",
    "repetir": "to repeat",
    "aprender": "to learn",
    "comprar": "to buy",
    "subir": "to go up",
    "evitar": "to avoid",
    "interesar": "to interest",
    "cerrar": "to close",
    "coger": "to take / catch",
    "responder": "to answer",
    "sufrir": "to suffer",
    "obtener": "to obtain",
    "observar": "to observe",
    "indicar": "to indicate",
    "imaginar": "to imagine",
    "soler": "to usually do",
    "actuar": "to act",
    "poner": "to put",
    "salir": "to leave / go out",
    "dar": "to give",
    "volver": "to return",
    "encontrar": "to find",
    "seguir": "to follow / continue",
    "agregar": "to add",
    "agua": "water",
    "mano": "hand",
    "país": "country",
    "ciudad": "city",
    "persona": "person",
    "niño": "boy / child",
    "niña": "girl",
    "amiga": "friend (f.)",
    "familia": "family",
    "hijo": "son",
    "hija": "daughter",
    "hermano": "brother",
    "hermana": "sister",
    "abuelo": "grandfather",
    "abuela": "grandmother",
    "perro": "dog",
    "gato": "cat",
    "libro": "book",
    "escuela": "school",
    "comida": "food",
    "coche": "car",
    "carro": "car",
    "calle": "street",
    "puerta": "door",
    "ventana": "window",
    "mesa": "table",
    "silla": "chair",
    "cama": "bed",
    "cuerpo": "body",
    "cabeza": "head",
    "ojo": "eye",
    "boca": "mouth",
    "corazón": "heart",
    "amor": "love",
    "guerra": "war",
    "paz": "peace",
    "luz": "light",
    "sol": "sun",
    "luna": "moon",
    "mar": "sea",
    "río": "river",
    "fuego": "fire",
    "tierra": "earth / land",
    "aire": "air",
    "color": "color",
    "número": "number",
    "palabra": "word",
    "historia": "history / story",
    "música": "music",
    "arte": "art",
    "idea": "idea",
    "forma": "form / way",
    "manera": "way",
    "ejemplo": "example",
    "razón": "reason",
    "caso": "case",
    "grupo": "group",
    "sistema": "system",
    "nivel": "level",
    "punto": "point",
    "lado": "side",
    "fin": "end",
    "final": "final / end",
    "principio": "beginning",
    "medio": "middle / means",
    "centro": "center",
    "norte": "north",
    "sur": "south",
    "oeste": "west",
    "izquierda": "left",
    "derecha": "right",
    "frente": "front",
    "fondo": "bottom / back",
    "alto": "tall / high",
    "bajo": "short / low",
    "largo": "long",
    "corto": "short",
    "pequeño": "small",
    "joven": "young",
    "viejo": "old",
    "vieja": "old (f.)",
    "bonito": "pretty",
    "feo": "ugly",
    "fácil": "easy",
    "difícil": "difficult",
    "importante": "important",
    "posible": "possible",
    "necesario": "necessary",
    "diferente": "different",
    "igual": "equal / same",
    "único": "only / unique",
    "próximo": "next",
    "último": "last",
    "primero": "first",
    "segundo": "second",
    "tercero": "third",
    "malo": "bad",
    "feliz": "happy",
    "triste": "sad",
    "cansado": "tired",
    "ocupado": "busy",
    "listo": "ready / smart",
    "rico": "rich / tasty",
    "pobre": "poor",
    "rojo": "red",
    "azul": "blue",
    "verde": "green",
    "negro": "black",
    "blanco": "white",
    "amarillo": "yellow",
    "gris": "gray",
    "lunes": "Monday",
    "martes": "Tuesday",
    "miércoles": "Wednesday",
    "jueves": "Thursday",
    "viernes": "Friday",
    "sábado": "Saturday",
    "domingo": "Sunday",
    "enero": "January",
    "febrero": "February",
    "marzo": "March",
    "abril": "April",
    "mayo": "May",
    "junio": "June",
    "julio": "July",
    "agosto": "August",
    "septiembre": "September",
    "octubre": "October",
    "noviembre": "November",
    "diciembre": "December",
    "cuatro": "four",
    "cinco": "five",
    "seis": "six",
    "siete": "seven",
    "ocho": "eight",
    "nueve": "nine",
    "diez": "ten",
    "cien": "hundred",
    "mil": "thousand",
    "ayer": "yesterday",
    "siempre": "always",
    "cerca": "near",
    "lejos": "far",
    "dentro": "inside",
    "fuera": "outside",
    "arriba": "up",
    "abajo": "down",
    "rápido": "fast",
    "lento": "slow",
    "temprano": "early",
    "pronto": "soon",
    "todavía": "still",
    "aún": "still / even",
    "quizá": "maybe",
    "cada": "each",
    "cualquier": "any",
    "alguno": "some",
    "ninguno": "none",
    "propio": "own",
    "cierto": "certain / true",
    "verdadero": "true",
    "falso": "false",
    "real": "real",
    "completo": "complete",
    "simple": "simple",
    "especial": "special",
    "general": "general",
    "social": "social",
    "político": "political",
    "económico": "economic",
    "público": "public",
    "privado": "private",
    "nacional": "national",
    "internacional": "international",
    "personal": "personal",
    "carrera": "career / race",
    "coser": "to sew",
    "mierda": "shit (vulgar)",
}

# ---------------------------------------------------------------------------
# Clean verb meanings for Verbs_Present sheet (infinitive form only)
# ---------------------------------------------------------------------------
VERB_MEANINGS: dict[str, str] = {
    "ser": "to be",
    "estar": "to be",
    "haber": "to have (auxiliary)",
    "tener": "to have",
    "ir": "to go",
    "venir": "to come",
    "hacer": "to do/make",
    "poder": "to be able to",
    "querer": "to want/love",
    "decir": "to say/tell",
    "dar": "to give",
    "ver": "to see",
    "saber": "to know",
    "poner": "to put",
    "salir": "to leave/go out",
    "traer": "to bring",
    "oír": "to hear",
    "caer": "to fall",
    "conocer": "to know (people/places)",
    "parecer": "to seem",
    "traducir": "to translate",
    "producir": "to produce",
    "seguir": "to follow/continue",
    "conseguir": "to get/achieve",
    "pedir": "to ask for",
    "servir": "to serve",
    "repetir": "to repeat",
    "reír": "to laugh",
    "sonreír": "to smile",
    "dormir": "to sleep",
    "morir": "to die",
    "volver": "to return",
    "encontrar": "to find",
    "recordar": "to remember",
    "contar": "to count/tell",
    "costar": "to cost",
    "mostrar": "to show",
    "jugar": "to play",
    "pensar": "to think",
    "cerrar": "to close",
    "empezar": "to begin",
    "comenzar": "to begin",
    "entender": "to understand",
    "perder": "to lose",
    "preferir": "to prefer",
    "sentir": "to feel",
    "mentir": "to lie",
    "divertir": "to entertain",
    "sugerir": "to suggest",
    "huir": "to flee",
    "construir": "to build",
    "incluir": "to include",
    "valer": "to be worth",
    "satisfacer": "to satisfy",
    "caber": "to fit",
    "andar": "to walk",
    "casar": "to marry",
    "coser": "to sew",
    "hablar": "to speak",
    "trabajar": "to work",
    "mirar": "to look/watch",
    "esperar": "to wait/hope",
    "pasar": "to pass/happen",
    "dejar": "to leave/let",
    "llamar": "to call",
    "llevar": "to carry/wear",
    "llegar": "to arrive",
    "tomar": "to take/drink",
    "buscar": "to look for",
    "entrar": "to enter",
    "escribir": "to write",
    "vivir": "to live",
    "existir": "to exist",
    "ocurrir": "to occur",
    "recibir": "to receive",
    "permitir": "to allow",
    "decidir": "to decide",
    "subir": "to go up",
    "abrir": "to open",
    "cubrir": "to cover",
    "descubrir": "to discover",
    "sufrir": "to suffer",
    "compartir": "to share",
    "partir": "to leave/divide",
    "leer": "to read",
    "creer": "to believe",
    "poseer": "to possess",
    "comer": "to eat",
    "beber": "to drink",
    "correr": "to run",
    "aprender": "to learn",
    "comprender": "to understand",
    "responder": "to respond",
    "vender": "to sell",
    "romper": "to break",
    "establecer": "to establish",
    "pertenecer": "to belong",
    "agradecer": "to thank",
    "gustar": "to like",
    "importar": "to matter",
    "interesar": "to interest",
    "faltar": "to lack/miss",
    "quedar": "to remain/meet",
    "necesitar": "to need",
    "utilizar": "to use",
    "usar": "to use",
    "estudiar": "to study",
    "enseñar": "to teach",
    "explicar": "to explain",
    "preguntar": "to ask",
    "ayudar": "to help",
    "escuchar": "to listen",
    "tocar": "to touch/play",
    "sacar": "to take out",
    "pagar": "to pay",
    "ganar": "to win/earn",
    "cambiar": "to change",
    "terminar": "to finish",
    "acabar": "to finish",
    "continuar": "to continue",
    "intentar": "to attempt",
    "tratar": "to treat/try",
    "lograr": "to achieve",
    "evitar": "to avoid",
    "aceptar": "to accept",
    "presentar": "to present",
    "considerar": "to consider",
    "realizar": "to carry out",
    "formar": "to form",
    "crear": "to create",
    "mantener": "to maintain",
    "obtener": "to obtain",
    "observar": "to observe",
    "indicar": "to indicate",
    "imaginar": "to imagine",
    "actuar": "to act",
    "soler": "to usually do",
    "suponer": "to suppose",
    "comprar": "to buy",
    "devolver": "to return (something)",
    "olvidar": "to forget",
    "levantar": "to raise",
    "sentar": "to seat",
    "acostar": "to put to bed",
    "despertar": "to wake up",
    "bañar": "to bathe",
    "duchar": "to shower",
    "lavar": "to wash",
    "limpiar": "to clean",
    "cocinar": "to cook",
    "viajar": "to travel",
    "caminar": "to walk",
    "nadar": "to swim",
    "bailar": "to dance",
    "cantar": "to sing",
    "llorar": "to cry",
    "gritar": "to yell",
    "besar": "to kiss",
    "abrazar": "to hug",
    "amar": "to love",
    "odiar": "to hate",
    "desear": "to desire",
    "temer": "to fear",
    "dudar": "to doubt",
    "descansar": "to rest",
    "soñar": "to dream",
    "contestar": "to answer",
    "bajar": "to go down",
    "enviar": "to send",
    "oler": "to smell",
    "probar": "to try/test",
    "envejecer": "to age",
    "divorciar": "to divorce",
    "nacer": "to be born",
    "crecer": "to grow",
    "obedecer": "to obey",
    "aparecer": "to appear",
    "ofrecer": "to offer",
    "reconocer": "to recognize",
    "dirigir": "to direct",
    "cumplir": "to fulfill",
    "alcanzar": "to reach",
    "convertir": "to convert",
    "resultar": "to result/turn out",
    "agregar": "to add",
    "deber": "should / must / to owe",
    "matar": "to kill",
    "funcionar": "to work / to function",
    "preocupar": "to worry",
    "significar": "to mean",
    "suceder": "to happen",
    "regresar": "to return",
    "encantar": "to delight / love (gustar-type)",
    "callar": "to be quiet",
    "disculpar": "to excuse / to apologize",
    "bastar": "to be enough",
    "jurar": "to swear",
    "prometer": "to promise",
    "sonar": "to sound / to ring",
    "salvar": "to save",
    "acercar": "to approach / to bring closer",
    "dañar": "to damage",
    "extrañar": "to miss / to find strange",
}

# ---------------------------------------------------------------------------
# Clean noun/adjective meanings for Nouns_Adjectives sheet
# ---------------------------------------------------------------------------
NOUN_ADJ_MEANINGS: dict[str, str] = {
    # Nouns
    "casa": "house/home",
    "vida": "life",
    "noche": "night",
    "vez": "time (occasion)",
    "cosa": "thing",
    "gente": "people",
    "mujer": "woman",
    "madre": "mother",
    "mamá": "mom",
    "hermana": "sister",
    "hija": "daughter",
    "amiga": "friend (f.)",
    "niña": "girl",
    "abuela": "grandmother",
    "mano": "hand",
    "agua": "water",
    "ciudad": "city",
    "familia": "family",
    "escuela": "school",
    "comida": "food",
    "calle": "street",
    "puerta": "door",
    "ventana": "window",
    "mesa": "table",
    "silla": "chair",
    "cama": "bed",
    "cabeza": "head",
    "boca": "mouth",
    "luz": "light",
    "luna": "moon",
    "paz": "peace",
    "guerra": "war",
    "palabra": "word",
    "historia": "history/story",
    "música": "music",
    "idea": "idea",
    "forma": "form/way",
    "manera": "way",
    "razón": "reason",
    "hora": "hour",
    "mañana": "morning/tomorrow",
    "tarde": "afternoon/late",
    "persona": "person",
    "verdad": "truth",
    "parte": "part",
    "tierra": "earth/land",
    "hombre": "man",
    "padre": "father",
    "hijo": "son",
    "hermano": "brother",
    "amigo": "friend (m.)",
    "niño": "boy/child",
    "abuelo": "grandfather",
    "papá": "dad",
    "señor": "sir/Mr.",
    "dios": "god",
    "día": "day",
    "año": "year",
    "tiempo": "time/weather",
    "mundo": "world",
    "lugar": "place",
    "trabajo": "work/job",
    "dinero": "money",
    "momento": "moment",
    "nombre": "name",
    "problema": "problem",
    "libro": "book",
    "perro": "dog",
    "gato": "cat",
    "coche": "car",
    "carro": "car",
    "cuerpo": "body",
    "ojo": "eye",
    "corazón": "heart",
    "amor": "love",
    "sol": "sun",
    "mar": "sea",
    "río": "river",
    "fuego": "fire",
    "aire": "air",
    "color": "color",
    "número": "number",
    "grupo": "group",
    "sistema": "system",
    "nivel": "level",
    "punto": "point",
    "lado": "side",
    "fin": "end",
    "país": "country",
    "ejemplo": "example",
    "caso": "case",
    "estado": "state",
    "acuerdo": "agreement",
    "tipo": "type/guy",
    "favor": "favor",
    "carrera": "career/race",
    "lunes": "Monday",
    "martes": "Tuesday",
    "miércoles": "Wednesday",
    "jueves": "Thursday",
    "viernes": "Friday",
    "sábado": "Saturday",
    "domingo": "Sunday",
    "enero": "January",
    "febrero": "February",
    "marzo": "March",
    "abril": "April",
    "mayo": "May",
    "junio": "June",
    "julio": "July",
    "agosto": "August",
    "septiembre": "September",
    "octubre": "October",
    "noviembre": "November",
    "diciembre": "December",
    "mapa": "map",
    "tema": "theme/topic",
    "programa": "program",
    "arte": "art",
    "final": "final/end",
    "principio": "beginning",
    "medio": "middle/means",
    "centro": "center",
    "norte": "north",
    "sur": "south",
    "oeste": "west",
    "izquierda": "left",
    "derecha": "right",
    "frente": "front",
    "fondo": "bottom/back",
    "segundo": "second",
    "tercero": "third",
    "mierda": "shit (vulgar)",
    # Adjectives
    "bueno": "good",
    "malo": "bad",
    "nuevo": "new",
    "viejo": "old",
    "grande": "big",
    "pequeño": "small",
    "alto": "tall/high",
    "bajo": "short/low",
    "mucho": "much/a lot",
    "poco": "little/few",
    "todo": "all/every",
    "otro": "other",
    "mismo": "same",
    "primero": "first",
    "último": "last",
    "mejor": "better/best",
    "feliz": "happy",
    "triste": "sad",
    "fácil": "easy",
    "difícil": "difficult",
    "importante": "important",
    "posible": "possible",
    "solo": "alone/only",
    "tanto": "so much",
    "alguno": "some",
    "ninguno": "none",
    "cierto": "certain/true",
    "propio": "own",
    "único": "only/unique",
    "próximo": "next",
    "rojo": "red",
    "azul": "blue",
    "verde": "green",
    "negro": "black",
    "blanco": "white",
    "amarillo": "yellow",
    "gris": "gray",
    "rápido": "fast",
    "lento": "slow",
    "bonito": "pretty",
    "feo": "ugly",
    "rico": "rich/tasty",
    "pobre": "poor",
    "joven": "young",
    "seguro": "sure/safe",
    "claro": "clear",
    "necesario": "necessary",
    "diferente": "different",
    "igual": "equal/same",
    "completo": "complete",
    "especial": "special",
    "general": "general",
    "social": "social",
    "político": "political",
    "económico": "economic",
    "público": "public",
    "privado": "private",
    "nacional": "national",
    "internacional": "international",
    "personal": "personal",
    "verdadero": "true",
    "falso": "false",
    "real": "real",
    "simple": "simple",
    "cansado": "tired",
    "ocupado": "busy",
    "listo": "ready/smart",
    "vieja": "old (f.)",
    "gran": "great/big",
    "largo": "long",
    "corto": "short",
    "último": "last",
    "esposa": "wife",
    "esposo": "husband",
    "equipo": "team",
    "arma": "weapon",
    "culpa": "guilt / blame",
    "asesino": "murderer",
    "fuerza": "force / strength",
    "campo": "field",
    "placer": "pleasure",
    "negocio": "business",
    "pena": "sorrow / pity",
    "piso": "floor / apartment",
    "ataque": "attack",
    "pelo": "hair",
    "peligro": "danger",
    "oro": "gold",
    "cita": "appointment / date",
    "pelea": "fight",
    "calma": "calm",
    "cena": "dinner",
    "regalo": "gift",
    "disculpa": "apology / sorry",
    "dólar": "dollar",
    "par": "pair / even",
    "extraño": "strange",
    "sueño": "dream",
    "daño": "harm / damage",
    "duro": "hard / tough",
    "lamento": "lament",
    "llamada": "call",
    "asesina": "murderer (f.)",
}

# Noun gender: m or f (common)
GENDER: dict[str, str] = {
    "casa": "f",
    "vida": "f",
    "noche": "f",
    "vez": "f",
    "cosa": "f",
    "gente": "f",
    "mujer": "f",
    "mamá": "f",
    "hermana": "f",
    "hija": "f",
    "amiga": "f",
    "niña": "f",
    "abuela": "f",
    "mano": "f",
    "agua": "f",
    "ciudad": "f",
    "familia": "f",
    "escuela": "f",
    "comida": "f",
    "calle": "f",
    "puerta": "f",
    "ventana": "f",
    "mesa": "f",
    "silla": "f",
    "cama": "f",
    "cabeza": "f",
    "boca": "f",
    "luz": "f",
    "luna": "f",
    "paz": "f",
    "guerra": "f",
    "palabra": "f",
    "historia": "f",
    "música": "f",
    "idea": "f",
    "forma": "f",
    "manera": "f",
    "razón": "f",
    "hora": "f",
    "mañana": "f",
    "tarde": "f",
    "persona": "f",
    "verdad": "f",
    "parte": "f",
    "tierra": "f",
    "hombre": "m",
    "padre": "m",
    "hijo": "m",
    "hermano": "m",
    "amigo": "m",
    "niño": "m",
    "abuelo": "m",
    "papá": "m",
    "señor": "m",
    "dios": "m",
    "día": "m",
    "año": "m",
    "tiempo": "m",
    "mundo": "m",
    "lugar": "m",
    "trabajo": "m",
    "dinero": "m",
    "momento": "m",
    "nombre": "m",
    "problema": "m",
    "libro": "m",
    "perro": "m",
    "gato": "m",
    "coche": "m",
    "carro": "m",
    "cuerpo": "m",
    "ojo": "m",
    "corazón": "m",
    "amor": "m",
    "sol": "m",
    "mar": "m",
    "río": "m",
    "fuego": "m",
    "aire": "m",
    "color": "m",
    "número": "m",
    "grupo": "m",
    "sistema": "m",
    "nivel": "m",
    "punto": "m",
    "lado": "m",
    "fin": "m",
    "país": "m",
    "ejemplo": "m",
    "caso": "m",
    "estado": "m",
    "acuerdo": "m",
    "tipo": "m",
    "favor": "m",
    "solo": "m",
    "todo": "m",
    "poco": "m",
    "mucho": "m",
    "bueno": "m",
    "malo": "m",
    "nuevo": "m",
    "viejo": "m",
    "mismo": "m",
    "otro": "m",
    "primero": "m",
    "último": "m",
    "mejor": "m",
    "peor": "m",
    "grande": "mf",
    "pequeño": "m",
    "alto": "m",
    "bajo": "m",
    "feliz": "mf",
    "triste": "mf",
    "fácil": "mf",
    "difícil": "mf",
    "importante": "mf",
    "posible": "mf",
    "mapa": "m",
    "tema": "m",
    "programa": "m",
    "carrera": "f",
    "viernes": "m",
    "lunes": "m",
    "martes": "m",
    "miércoles": "m",
    "jueves": "m",
    "sábado": "m",
    "domingo": "m",
    "esposa": "f",
    "esposo": "m",
    "equipo": "m",
    "arma": "f",
    "culpa": "f",
    "asesino": "m",
    "fuerza": "f",
    "campo": "m",
    "placer": "m",
    "negocio": "m",
    "pena": "f",
    "piso": "m",
    "ataque": "m",
    "pelo": "m",
    "peligro": "m",
    "oro": "m",
    "cita": "f",
    "pelea": "f",
    "calma": "f",
    "cena": "f",
    "regalo": "m",
    "disculpa": "f",
    "dólar": "m",
    "par": "m",
    "sueño": "m",
    "daño": "m",
    "lamento": "m",
    "llamada": "f",
    "calle": "f",
}

# ---------------------------------------------------------------------------
# Verb conjugations (present indicative)
# ---------------------------------------------------------------------------
IRREGULAR: dict[str, dict[str, str]] = {
    "ser": {
        "yo": "soy",
        "tú": "eres",
        "él": "es",
        "nosotros": "somos",
        "vosotros": "sois",
        "ellos": "son",
    },
    "estar": {
        "yo": "estoy",
        "tú": "estás",
        "él": "está",
        "nosotros": "estamos",
        "vosotros": "estáis",
        "ellos": "están",
    },
    "haber": {
        "yo": "he",
        "tú": "has",
        "él": "ha",
        "nosotros": "hemos",
        "vosotros": "habéis",
        "ellos": "han",
    },
    "tener": {
        "yo": "tengo",
        "tú": "tienes",
        "él": "tiene",
        "nosotros": "tenemos",
        "vosotros": "tenéis",
        "ellos": "tienen",
    },
    "ir": {
        "yo": "voy",
        "tú": "vas",
        "él": "va",
        "nosotros": "vamos",
        "vosotros": "vais",
        "ellos": "van",
    },
    "venir": {
        "yo": "vengo",
        "tú": "vienes",
        "él": "viene",
        "nosotros": "venimos",
        "vosotros": "venís",
        "ellos": "vienen",
    },
    "hacer": {
        "yo": "hago",
        "tú": "haces",
        "él": "hace",
        "nosotros": "hacemos",
        "vosotros": "hacéis",
        "ellos": "hacen",
    },
    "poder": {
        "yo": "puedo",
        "tú": "puedes",
        "él": "puede",
        "nosotros": "podemos",
        "vosotros": "podéis",
        "ellos": "pueden",
    },
    "querer": {
        "yo": "quiero",
        "tú": "quieres",
        "él": "quiere",
        "nosotros": "queremos",
        "vosotros": "queréis",
        "ellos": "quieren",
    },
    "decir": {
        "yo": "digo",
        "tú": "dices",
        "él": "dice",
        "nosotros": "decimos",
        "vosotros": "decís",
        "ellos": "dicen",
    },
    "dar": {
        "yo": "doy",
        "tú": "das",
        "él": "da",
        "nosotros": "damos",
        "vosotros": "dais",
        "ellos": "dan",
    },
    "ver": {
        "yo": "veo",
        "tú": "ves",
        "él": "ve",
        "nosotros": "vemos",
        "vosotros": "veis",
        "ellos": "ven",
    },
    "saber": {
        "yo": "sé",
        "tú": "sabes",
        "él": "sabe",
        "nosotros": "sabemos",
        "vosotros": "sabéis",
        "ellos": "saben",
    },
    "poner": {
        "yo": "pongo",
        "tú": "pones",
        "él": "pone",
        "nosotros": "ponemos",
        "vosotros": "ponéis",
        "ellos": "ponen",
    },
    "salir": {
        "yo": "salgo",
        "tú": "sales",
        "él": "sale",
        "nosotros": "salimos",
        "vosotros": "salís",
        "ellos": "salen",
    },
    "traer": {
        "yo": "traigo",
        "tú": "traes",
        "él": "trae",
        "nosotros": "traemos",
        "vosotros": "traéis",
        "ellos": "traen",
    },
    "oír": {
        "yo": "oigo",
        "tú": "oyes",
        "él": "oye",
        "nosotros": "oímos",
        "vosotros": "oís",
        "ellos": "oyen",
    },
    "caer": {
        "yo": "caigo",
        "tú": "caes",
        "él": "cae",
        "nosotros": "caemos",
        "vosotros": "caéis",
        "ellos": "caen",
    },
    "conocer": {
        "yo": "conozco",
        "tú": "conoces",
        "él": "conoce",
        "nosotros": "conocemos",
        "vosotros": "conocéis",
        "ellos": "conocen",
    },
    "parecer": {
        "yo": "parezco",
        "tú": "pareces",
        "él": "parece",
        "nosotros": "parecemos",
        "vosotros": "parecéis",
        "ellos": "parecen",
    },
    "traducir": {
        "yo": "traduzco",
        "tú": "traduces",
        "él": "traduce",
        "nosotros": "traducimos",
        "vosotros": "traducís",
        "ellos": "traducen",
    },
    "producir": {
        "yo": "produzco",
        "tú": "produces",
        "él": "produce",
        "nosotros": "producimos",
        "vosotros": "producís",
        "ellos": "producen",
    },
    "seguir": {
        "yo": "sigo",
        "tú": "sigues",
        "él": "sigue",
        "nosotros": "seguimos",
        "vosotros": "seguís",
        "ellos": "siguen",
    },
    "conseguir": {
        "yo": "consigo",
        "tú": "consigues",
        "él": "consigue",
        "nosotros": "conseguimos",
        "vosotros": "conseguís",
        "ellos": "consiguen",
    },
    "pedir": {
        "yo": "pido",
        "tú": "pides",
        "él": "pide",
        "nosotros": "pedimos",
        "vosotros": "pedís",
        "ellos": "piden",
    },
    "servir": {
        "yo": "sirvo",
        "tú": "sirves",
        "él": "sirve",
        "nosotros": "servimos",
        "vosotros": "servís",
        "ellos": "sirven",
    },
    "repetir": {
        "yo": "repito",
        "tú": "repites",
        "él": "repite",
        "nosotros": "repetimos",
        "vosotros": "repetís",
        "ellos": "repiten",
    },
    "reír": {
        "yo": "río",
        "tú": "ríes",
        "él": "ríe",
        "nosotros": "reímos",
        "vosotros": "reís",
        "ellos": "ríen",
    },
    "sonreír": {
        "yo": "sonrío",
        "tú": "sonríes",
        "él": "sonríe",
        "nosotros": "sonreímos",
        "vosotros": "sonreís",
        "ellos": "sonríen",
    },
    "dormir": {
        "yo": "duermo",
        "tú": "duermes",
        "él": "duerme",
        "nosotros": "dormimos",
        "vosotros": "dormís",
        "ellos": "duermen",
    },
    "morir": {
        "yo": "muero",
        "tú": "mueres",
        "él": "muere",
        "nosotros": "morimos",
        "vosotros": "morís",
        "ellos": "mueren",
    },
    "volver": {
        "yo": "vuelvo",
        "tú": "vuelves",
        "él": "vuelve",
        "nosotros": "volvemos",
        "vosotros": "volvéis",
        "ellos": "vuelven",
    },
    "encontrar": {
        "yo": "encuentro",
        "tú": "encuentras",
        "él": "encuentra",
        "nosotros": "encontramos",
        "vosotros": "encontráis",
        "ellos": "encuentran",
    },
    "recordar": {
        "yo": "recuerdo",
        "tú": "recuerdas",
        "él": "recuerda",
        "nosotros": "recordamos",
        "vosotros": "recordáis",
        "ellos": "recuerdan",
    },
    "contar": {
        "yo": "cuento",
        "tú": "cuentas",
        "él": "cuenta",
        "nosotros": "contamos",
        "vosotros": "contáis",
        "ellos": "cuentan",
    },
    "costar": {
        "yo": "cuesto",
        "tú": "cuestas",
        "él": "cuesta",
        "nosotros": "costamos",
        "vosotros": "costáis",
        "ellos": "cuestan",
    },
    "mostrar": {
        "yo": "muestro",
        "tú": "muestras",
        "él": "muestra",
        "nosotros": "mostramos",
        "vosotros": "mostráis",
        "ellos": "muestran",
    },
    "jugar": {
        "yo": "juego",
        "tú": "juegas",
        "él": "juega",
        "nosotros": "jugamos",
        "vosotros": "jugáis",
        "ellos": "juegan",
    },
    "pensar": {
        "yo": "pienso",
        "tú": "piensas",
        "él": "piensa",
        "nosotros": "pensamos",
        "vosotros": "pensáis",
        "ellos": "piensan",
    },
    "cerrar": {
        "yo": "cierro",
        "tú": "cierras",
        "él": "cierra",
        "nosotros": "cerramos",
        "vosotros": "cerráis",
        "ellos": "cierran",
    },
    "empezar": {
        "yo": "empiezo",
        "tú": "empiezas",
        "él": "empieza",
        "nosotros": "empezamos",
        "vosotros": "empezáis",
        "ellos": "empiezan",
    },
    "comenzar": {
        "yo": "comienzo",
        "tú": "comienzas",
        "él": "comienza",
        "nosotros": "comenzamos",
        "vosotros": "comenzáis",
        "ellos": "comienzan",
    },
    "entender": {
        "yo": "entiendo",
        "tú": "entiendes",
        "él": "entiende",
        "nosotros": "entendemos",
        "vosotros": "entendéis",
        "ellos": "entienden",
    },
    "perder": {
        "yo": "pierdo",
        "tú": "pierdes",
        "él": "pierde",
        "nosotros": "perdemos",
        "vosotros": "perdéis",
        "ellos": "pierden",
    },
    "preferir": {
        "yo": "prefiero",
        "tú": "prefieres",
        "él": "prefiere",
        "nosotros": "preferimos",
        "vosotros": "preferís",
        "ellos": "prefieren",
    },
    "sentir": {
        "yo": "siento",
        "tú": "sientes",
        "él": "siente",
        "nosotros": "sentimos",
        "vosotros": "sentís",
        "ellos": "sienten",
    },
    "mentir": {
        "yo": "miento",
        "tú": "mientes",
        "él": "miente",
        "nosotros": "mentimos",
        "vosotros": "mentís",
        "ellos": "mienten",
    },
    "divertir": {
        "yo": "divierto",
        "tú": "diviertes",
        "él": "divierte",
        "nosotros": "divertimos",
        "vosotros": "divertís",
        "ellos": "divierten",
    },
    "sugerir": {
        "yo": "sugiero",
        "tú": "sugieres",
        "él": "sugiere",
        "nosotros": "sugerimos",
        "vosotros": "sugerís",
        "ellos": "sugieren",
    },
    "huir": {
        "yo": "huyo",
        "tú": "huyes",
        "él": "huye",
        "nosotros": "huimos",
        "vosotros": "huís",
        "ellos": "huyen",
    },
    "construir": {
        "yo": "construyo",
        "tú": "construyes",
        "él": "construye",
        "nosotros": "construimos",
        "vosotros": "construís",
        "ellos": "construyen",
    },
    "incluir": {
        "yo": "incluyo",
        "tú": "incluyes",
        "él": "incluye",
        "nosotros": "incluimos",
        "vosotros": "incluís",
        "ellos": "incluyen",
    },
    "valer": {
        "yo": "valgo",
        "tú": "vales",
        "él": "vale",
        "nosotros": "valemos",
        "vosotros": "valéis",
        "ellos": "valen",
    },
    "satisfacer": {
        "yo": "satisfago",
        "tú": "satisfaces",
        "él": "satisface",
        "nosotros": "satisfacemos",
        "vosotros": "satisfacéis",
        "ellos": "satisfacen",
    },
    "caber": {
        "yo": "quepo",
        "tú": "cabes",
        "él": "cabe",
        "nosotros": "cabemos",
        "vosotros": "cabéis",
        "ellos": "caben",
    },
    "andar": {
        "yo": "ando",
        "tú": "andas",
        "él": "anda",
        "nosotros": "andamos",
        "vosotros": "andáis",
        "ellos": "andan",
    },
}

# e→ie, o→ue, e→i stem changers (boot) for regular endings
E_IE = {
    "cerrar",
    "empezar",
    "comenzar",
    "pensar",
    "sentar",
    "negar",
    "atravesar",
    "calentar",
    "confesar",
    "defender",
    "encender",
    "entender",
    "perder",
    "querer",
    "ascender",
    "atender",
}
O_UE = {
    "contar",
    "costar",
    "encontrar",
    "mostrar",
    "recordar",
    "volar",
    "almorzar",
    "aprobar",
    "colgar",
    "probar",
    "soñar",
    "volver",
    "devolver",
    "envolver",
    "mover",
    "doler",
    "llover",
    "poder",
    "dormir",
    "morir",
    "resolver",
}
E_I = {
    "pedir",
    "servir",
    "repetir",
    "seguir",
    "conseguir",
    "vestir",
    "medir",
    "reír",
    "sonreír",
    "elegir",
    "corregir",
    "competir",
    "freír",
}
U_UE = {"jugar"}

NON_VERB_LOOKALIKES = {
    "lugar",
    "mujer",
    "señor",
    "mejor",
    "peor",
    "mar",
    "par",
    "favor",
    "color",
    "amor",
    "dolor",
    "calor",
    "actor",
    "motor",
    "taller",
    "mayor",
    "menor",
    "cualquier",
    "primer",
    "tercer",
    "carácter",
    "dólar",
    "ayer",
    "placer",
    "azúcar",
    "altar",
    "collar",
    "celular",
    "familiar",
    "popular",
    "regular",
    "similar",
    "particular",
    "militar",
    "solar",
    "polar",
    "radar",
    "máster",
    "líder",
}

PRONOUNS = {
    "yo",
    "tú",
    "él",
    "ella",
    "usted",
    "nosotros",
    "nosotras",
    "vosotros",
    "ellos",
    "ellas",
    "ustedes",
    "me",
    "te",
    "se",
    "nos",
    "os",
    "le",
    "les",
    "lo",
    "la",
    "los",
    "las",
    "mi",
    "tu",
    "su",
    "mis",
    "tus",
    "sus",
    "mí",
    "ti",
    "conmigo",
    "contigo",
    "mío",
    "nuestro",
    "esto",
    "eso",
    "este",
    "esta",
    "estos",
    "estas",
    "ese",
    "esa",
    "esos",
    "esas",
    "aquel",
    "aquella",
    "aquellos",
    "aquellas",
    "aquello",
    "ello",
    "alguien",
    "nadie",
}

PREPOSITIONS = {
    "de",
    "a",
    "en",
    "con",
    "por",
    "para",
    "sin",
    "sobre",
    "hasta",
    "desde",
    "hacia",
    "entre",
    "durante",
    "según",
    "contra",
    "del",
    "al",
    "acerca",
}

CONJUNCTIONS = {
    "y",
    "o",
    "pero",
    "porque",
    "si",
    "que",
    "como",
    "cuando",
    "donde",
    "ni",
    "mas",
    "aunque",
    "sino",
    "mientras",
}

ADVERBS = {
    "no",
    "sí",
    "muy",
    "más",
    "menos",
    "ya",
    "aún",
    "también",
    "tampoco",
    "nunca",
    "siempre",
    "ahora",
    "antes",
    "después",
    "luego",
    "aquí",
    "allí",
    "ahí",
    "bien",
    "mal",
    "así",
    "solo",
    "sólo",
    "quizá",
    "todavía",
    "pronto",
    "cerca",
    "lejos",
    "dentro",
    "fuera",
    "arriba",
    "abajo",
    "hoy",
    "ayer",
    "mañana",
    "entonces",
    "tan",
    "mucho",
    "poco",
    "nada",
    "algo",
}

ARTICLES = {"el", "la", "los", "las", "un", "una", "unos", "unas", "lo"}

QUESTION_WORDS = {
    "qué",
    "quién",
    "dónde",
    "cuándo",
    "cómo",
    "cuál",
    "cuánto",
    "por qué",
}

ADJECTIVES = {
    "bueno",
    "malo",
    "nuevo",
    "viejo",
    "grande",
    "pequeño",
    "alto",
    "bajo",
    "mucho",
    "poco",
    "todo",
    "otro",
    "mismo",
    "primero",
    "último",
    "mejor",
    "peor",
    "mayor",
    "menor",
    "feliz",
    "triste",
    "fácil",
    "difícil",
    "importante",
    "posible",
    "solo",
    "tanto",
    "alguno",
    "ninguno",
    "cierto",
    "propio",
    "único",
    "próximo",
    "rojo",
    "azul",
    "verde",
    "negro",
    "blanco",
    "amarillo",
    "gris",
    "rápido",
    "lento",
    "bonito",
    "feo",
    "rico",
    "pobre",
    "joven",
    "seguro",
    "claro",
    "necesario",
    "diferente",
    "igual",
    "completo",
    "especial",
    "extraño",
    "duro",
    "largo",
    "corto",
    "gran",
    "cualquier",
    "primer",
    "tercer",
}


def _norm_word(word: str) -> str:
    return (word or "").strip().lower()


def _lexicon_nonverb_type_exact(w: str) -> str | None:
    if w in ARTICLES:
        return "article"
    if w in QUESTION_WORDS:
        return "question word"
    if w in PRONOUNS:
        return "pronoun"
    if w in PREPOSITIONS:
        return "preposition"
    if w in CONJUNCTIONS:
        return "conjunction"
    if w in ADVERBS:
        return "adverb"
    if w in ADJECTIVES:
        return "adjective"
    if w in NON_VERB_LOOKALIKES:
        return "adjective" if w in ADJECTIVES else "noun"
    if w in NOUN_ADJ_MEANINGS or w in GENDER:
        return "adjective" if w in ADJECTIVES else "noun"
    return None


def lexicon_nonverb_type(word: str) -> str | None:
    """POS for a known non-verb. Used so noun/adj/adv surfaces are not conjugated."""
    w = _norm_word(word)
    if not w:
        return None
    candidates = [w]
    if w.endswith("es") and len(w) > 3:
        candidates.append(w[:-2])
    if w.endswith("s") and len(w) > 2:
        candidates.append(w[:-1])
    for candidate in candidates:
        pos = _lexicon_nonverb_type_exact(candidate)
        if pos:
            return pos
    return None


def is_verb(lemma: str) -> bool:
    if lemma in IRREGULAR:
        return True
    if lemma in NON_VERB_LOOKALIKES:
        return False
    if lemma in {
        "ser",
        "estar",
        "haber",
        "ir",
        "ver",
        "dar",
        "oír",
        "reír",
        "freír",
        "poder",
        "querer",
        "tener",
        "venir",
        "hacer",
        "decir",
        "saber",
        "poner",
        "salir",
        "traer",
        "caer",
        "valer",
        "caber",
        "andar",
        "jugar",
        "pensar",
        "contar",
        "volver",
        "dormir",
        "pedir",
        "seguir",
        "sentir",
        "preferir",
        "encontrar",
        "recordar",
        "empezar",
        "comenzar",
        "entender",
        "perder",
        "cerrar",
        "mostrar",
        "costar",
        "almorzar",
        "mover",
        "doler",
        "llover",
        "morir",
        "vestir",
        "medir",
        "elegir",
        "conseguir",
        "sonreír",
        "incluir",
        "construir",
        "huir",
        "traducir",
        "producir",
        "conocer",
        "parecer",
        "aparecer",
        "ofrecer",
        "nacer",
        "crecer",
        "obedecer",
        "casar",
        "coser",
        "hablar",
        "trabajar",
        "mirar",
        "esperar",
        "pasar",
        "dejar",
        "llamar",
        "llevar",
        "llegar",
        "tomar",
        "buscar",
        "entrar",
        "escribir",
        "vivir",
        "existir",
        "ocurrir",
        "recibir",
        "permitir",
        "decidir",
        "subir",
        "abrir",
        "cubrir",
        "descubrir",
        "sufrir",
        "compartir",
        "partir",
        "leer",
        "creer",
        "poseer",
        "comer",
        "beber",
        "correr",
        "aprender",
        "comprender",
        "responder",
        "vender",
        "romper",
        "establecer",
        "pertenecer",
        "agradecer",
        "gustar",
        "importar",
        "interesar",
        "faltar",
        "quedar",
        "necesitar",
        "utilizar",
        "usar",
        "estudiar",
        "enseñar",
        "explicar",
        "preguntar",
        "ayudar",
        "escuchar",
        "tocar",
        "sacar",
        "pagar",
        "ganar",
        "cambiar",
        "terminar",
        "acabar",
        "continuar",
        "intentar",
        "tratar",
        "lograr",
        "evitar",
        "aceptar",
        "presentar",
        "considerar",
        "realizar",
        "formar",
        "crear",
        "mantener",
        "obtener",
        "observar",
        "indicar",
        "imaginar",
        "actuar",
        "soler",
        "suponer",
        "comprar",
        "devolver",
        "olvidar",
        "levantar",
        "sentar",
        "acostar",
        "despertar",
        "bañar",
        "duchar",
        "lavar",
        "limpiar",
        "cocinar",
        "viajar",
        "caminar",
        "nadar",
        "bailar",
        "cantar",
        "llorar",
        "gritar",
        "besar",
        "abrazar",
        "amar",
        "odiar",
        "desear",
        "temer",
        "dudar",
        "descansar",
        "soñar",
        "contestar",
        "bajar",
        "enviar",
        "oler",
        "probar",
        "envejecer",
        "divorciar",
    }:
        return True
    if lexicon_nonverb_type(lemma):
        return False
    if re.search(r"(ar|er|ir|ír)$", lemma) and len(lemma) > 3:
        return lemma not in NON_VERB_LOOKALIKES
    return False


def conjugate_regular(inf: str) -> dict[str, str] | None:
    if inf in IRREGULAR:
        return IRREGULAR[inf]
    if inf in NON_VERB_LOOKALIKES or lexicon_nonverb_type(inf):
        return None
    if not re.search(r"(ar|er|ir|ír)$", inf):
        return None

    stem, ending = inf[:-2], inf[-2:]

    # Determine if this verb has a stem change
    boot_stem = stem

    if inf in E_IE:
        # Find last 'e' and change to 'ie'
        idx = stem.rfind("e")
        if idx != -1:
            boot_stem = stem[:idx] + "ie" + stem[idx + 1 :]
    elif inf in O_UE:
        idx = stem.rfind("o")
        if idx != -1:
            boot_stem = stem[:idx] + "ue" + stem[idx + 1 :]
    elif inf in E_I:
        idx = stem.rfind("e")
        if idx != -1:
            boot_stem = stem[:idx] + "i" + stem[idx + 1 :]
    elif inf in U_UE:
        idx = stem.rfind("u")
        if idx != -1:
            boot_stem = stem[:idx] + "ue" + stem[idx + 1 :]

    if ending == "ar":
        return {
            "yo": boot_stem + "o",
            "tú": boot_stem + "as",
            "él": boot_stem + "a",
            "nosotros": stem + "amos",
            "vosotros": stem + "áis",
            "ellos": boot_stem + "an",
        }
    if ending == "er":
        return {
            "yo": boot_stem + "o",
            "tú": boot_stem + "es",
            "él": boot_stem + "e",
            "nosotros": stem + "emos",
            "vosotros": stem + "éis",
            "ellos": boot_stem + "en",
        }
    if ending in ("ir", "ír"):
        return {
            "yo": boot_stem + "o",
            "tú": boot_stem + "es",
            "él": boot_stem + "e",
            "nosotros": stem + "imos",
            "vosotros": stem + "ís",
            "ellos": boot_stem + "en",
        }
    return None


def guess_type(lemma: str, surface: str) -> str:
    # Trust the written word when we know it (casa ≠ casar, dólares ≠ a verb).
    surface_pos = lexicon_nonverb_type(surface)
    if surface_pos:
        return surface_pos
    lemma_pos = lexicon_nonverb_type(lemma)
    if lemma_pos and lemma not in IRREGULAR and lemma not in VERB_MEANINGS:
        return lemma_pos
    if lemma in IRREGULAR or is_verb(lemma):
        if lemma in NON_VERB_LOOKALIKES:
            return "adjective" if lemma in ADJECTIVES else "noun"
        if conjugate_regular(lemma) or lemma in IRREGULAR:
            return "verb"
    if lemma in GENDER or surface in GENDER:
        g = GENDER.get(lemma) or GENDER.get(surface)
        if g in ("m", "f", "mf"):
            if lemma in ADJECTIVES or surface in ADJECTIVES:
                return "adjective"
            return "noun"
    if lemma in PRONOUNS or surface in PRONOUNS:
        return "pronoun"
    if lemma in PREPOSITIONS or surface in PREPOSITIONS:
        return "preposition"
    if lemma in CONJUNCTIONS or surface in CONJUNCTIONS:
        return "conjunction"
    if lemma in ADVERBS or surface in ADVERBS:
        return "adverb"
    if lemma in ARTICLES or surface in ARTICLES:
        return "article"
    if lemma in QUESTION_WORDS or surface in QUESTION_WORDS:
        return "question word"
    if re.search(r"(ar|er|ir)$", lemma) and conjugate_regular(lemma):
        return "verb"
    if (
        lemma.endswith("o")
        or lemma.endswith("or")
        or lemma.endswith("aje")
        or lemma.endswith("án")
    ):
        return "noun"
    if (
        lemma.endswith("a")
        or lemma.endswith("ción")
        or lemma.endswith("sión")
        or lemma.endswith("dad")
    ):
        return "noun"
    if lemma.endswith("mente"):
        return "adverb"
    return "other"


def english_meaning(lemma: str, surface: str = "", typ: str = "") -> str:
    """Look up an English gloss. Never fall back to the Spanish word itself."""
    keys: list[str] = []
    for raw in (lemma, surface):
        if not raw:
            continue
        text = raw.strip()
        if text and text not in keys:
            keys.append(text)
        low = text.lower()
        if low and low not in keys:
            keys.append(low)

    if typ == "verb":
        for key in keys:
            if key in VERB_MEANINGS:
                return VERB_MEANINGS[key]
        for key in keys:
            val = MEANINGS.get(key, "")
            if val.lower().startswith("to ") or " / to " in val.lower():
                return val
        return ""
    if typ in ("noun", "adjective"):
        dicts = (NOUN_ADJ_MEANINGS, MEANINGS)
    else:
        dicts = (MEANINGS, NOUN_ADJ_MEANINGS)

    for d in dicts:
        for key in keys:
            val = d.get(key)
            if val:
                return val
    return ""


# ---------------------------------------------------------------------------
# Pronunciation (English-friendly, LatAm-leaning)
# ---------------------------------------------------------------------------
def pronounce(word: str) -> str:
    if not word or word == "—":
        return "—"
    w = word.lower().strip()
    OV = {
        "que": "keh",
        "de": "deh",
        "no": "noh",
        "a": "ah",
        "la": "lah",
        "el": "ehl",
        "y": "ee",
        "en": "ehn",
        "lo": "loh",
        "un": "oon",
        "por": "por",
        "qué": "keh",
        "me": "meh",
        "te": "teh",
        "se": "seh",
        "con": "kohn",
        "para": "PAH-rah",
        "mi": "mee",
        "si": "see",
        "sí": "SEE",
        "su": "soo",
        "tu": "too",
        "tú": "too",
        "yo": "yoh",
        "él": "ehl",
        "ella": "EH-yah",
        "muy": "mwee",
        "hay": "eye",
        "hoy": "oy",
        "soy": "soy",
        "voy": "boy",
        "doy": "doy",
        "es": "ehs",
        "son": "sohn",
        "ser": "sehr",
        "ir": "eer",
        "ver": "behr",
        "dar": "dahr",
        "hola": "OH-lah",
        "gracias": "GRAH-syahs",
        "bueno": "BWEH-noh",
        "buena": "BWEH-nah",
        "casa": "KAH-sah",
        "agua": "AH-gwah",
        "tiempo": "TYEHM-poh",
        "también": "tahm-BYEHN",
        "después": "dehs-PWEHS",
        "aquí": "ah-KEE",
        "allí": "ah-YEE",
        "ahí": "ah-EE",
        "cómo": "KOH-moh",
        "dónde": "DOHN-deh",
        "cuándo": "KWAHN-doh",
        "quién": "kyehn",
        "cuál": "kwahl",
        "cuánto": "KWAHN-toh",
        "nosotros": "noh-SOH-trohs",
        "vosotros": "boh-SOH-trohs",
        "ustedes": "oos-TEH-dehs",
        "usted": "oos-TEHD",
        "ellos": "EH-yohs",
        "ellas": "EH-yahs",
        "esto": "EHS-toh",
        "eso": "EHS-soh",
        "este": "EHS-teh",
        "esta": "EHS-tah",
        "ese": "EHS-seh",
        "esa": "EHS-sah",
        "todo": "TOH-doh",
        "nada": "NAH-dah",
        "algo": "AHL-goh",
        "alguien": "ahl-GYEHN",
        "nadie": "NAH-dyeh",
        "siempre": "SYEHM-preh",
        "nunca": "NOON-kah",
        "ahora": "ah-OH-rah",
        "entonces": "ehn-TOHN-sehs",
        "porque": "POR-keh",
        "cuando": "KWAHN-doh",
        "donde": "DOHN-deh",
        "como": "KOH-moh",
        "más": "mahs",
        "menos": "MEH-nohs",
        "bien": "byehn",
        "mal": "mahl",
        "ya": "yah",
        "aún": "ah-OON",
        "solo": "SOH-loh",
        "sólo": "SOH-loh",
        "tan": "tahn",
        "tanto": "TAHN-toh",
        "mucho": "MOO-choh",
        "poco": "POH-koh",
        "cada": "KAH-dah",
        "del": "dehl",
        "al": "ahl",
        "mí": "mee",
        "ti": "tee",
        "uno": "OO-noh",
        "una": "OO-nah",
        "dos": "dohs",
        "tres": "trehs",
        "hacer": "ah-SEHR",
        "tener": "teh-NEHR",
        "poder": "poh-DEHR",
        "querer": "keh-REHR",
        "decir": "deh-SEER",
        "saber": "sah-BEHR",
        "poner": "poh-NEHR",
        "salir": "sah-LEER",
        "venir": "beh-NEHR",
        "haber": "ah-BEHR",
        "estar": "ehs-TAHR",
        "eres": "EH-rehs",
        "somos": "SOH-mohs",
        "sois": "soy-ees",
        "estoy": "ehs-TOY",
        "estás": "ehs-TAHS",
        "está": "ehs-TAH",
        "estamos": "ehs-TAH-mohs",
        "estáis": "ehs-TAH-ees",
        "están": "ehs-TAHN",
        "tengo": "TEHN-goh",
        "tienes": "TYEH-nehs",
        "tiene": "TYEH-neh",
        "tenemos": "teh-NEH-mohs",
        "tenéis": "teh-NEH-ees",
        "tienen": "TYEH-nehn",
        "vas": "bahs",
        "va": "bah",
        "vamos": "BAH-mohs",
        "vais": "bah-ees",
        "van": "bahn",
        "puedo": "PWEH-doh",
        "puedes": "PWEH-dehs",
        "puede": "PWEH-deh",
        "podemos": "poh-DEH-mohs",
        "pueden": "PWEH-dehn",
        "quiero": "KYEH-roh",
        "quieres": "KYEH-rehs",
        "quiere": "KYEH-reh",
        "hago": "AH-goh",
        "haces": "AH-sehs",
        "hace": "AH-seh",
        "hacemos": "ah-SEH-mohs",
        "hacen": "AH-sehn",
        "digo": "DEE-goh",
        "dices": "DEE-sehs",
        "dice": "DEE-seh",
        "decimos": "deh-SEE-mohs",
        "dicen": "DEE-sehn",
        "sé": "seh",
        "sabes": "SAH-behs",
        "sabe": "SAH-beh",
        "veo": "BEH-oh",
        "ves": "behs",
        "ve": "beh",
        "das": "dahs",
        "da": "dah",
        "he": "eh",
        "has": "ahs",
        "ha": "ah",
        "hemos": "EH-mohs",
        "han": "ahn",
        "vengo": "BEHN-goh",
        "vienes": "BYEH-nehs",
        "viene": "BYEH-neh",
        "adiós": "ah-DYOHS",
        "favor": "fah-BOHR",
        "hablo": "AH-bloh",
        "hablas": "AH-blahs",
        "habla": "AH-blah",
        "hablamos": "ah-BLAH-mohs",
        "hablan": "AH-blahn",
    }
    if w in OV:
        return OV[w]
    return approx_pron(word)


def approx_pron(word: str) -> str:
    """Rough English-friendly pronunciation."""
    w = word.lower()
    # restore common whole forms
    special = {
        "hola": "OH-lah",
        "gracias": "GRAH-syahs",
        "bueno": "BWEH-noh",
        "agua": "AH-gwah",
        "español": "ehs-pah-NYOHL",
        "inglés": "een-GLEHS",
        "mañana": "mah-NYAH-nah",
        "niño": "NEE-nyoh",
        "señor": "seh-NYOHR",
        "señora": "seh-NYOH-rah",
    }
    if w in special:
        return special[w]

    chars = list(w)
    out = []
    i = 0
    while i < len(chars):
        c = chars[i]
        nxt = chars[i + 1] if i + 1 < len(chars) else ""
        pair = c + nxt
        if pair == "ch":
            out.append("ch")
            i += 2
            continue
        if pair == "ll":
            out.append("y")
            i += 2
            continue
        if pair == "rr":
            out.append("rr")
            i += 2
            continue
        if pair in ("qu",) and nxt:
            out.append("k")
            i += 2
            continue
        if c == "q" and nxt == "u":
            out.append("k")
            i += 2
            continue
        if c == "g" and nxt == "u" and i + 2 < len(chars) and chars[i + 2] in "eiéí":
            out.append("g")
            i += 2
            continue
        if c == "g" and nxt in "eiéí":
            out.append("h")
            i += 1
            continue
        if c == "c" and nxt in "eiéí":
            out.append("s")
            i += 1
            continue
        if c == "c":
            out.append("k")
            i += 1
            continue
        if c == "j":
            out.append("h")
            i += 1
            continue
        if c == "ñ":
            out.append("ny")
            i += 1
            continue
        if c == "h":
            i += 1
            continue
        if c == "v":
            out.append("b")
            i += 1
            continue
        if c == "z":
            out.append("s")
            i += 1
            continue
        if c == "x":
            out.append("ks")
            i += 1
            continue
        if c == "á":
            out.append("AH")
            i += 1
            continue
        if c == "é":
            out.append("EH")
            i += 1
            continue
        if c == "í":
            out.append("EE")
            i += 1
            continue
        if c == "ó":
            out.append("OH")
            i += 1
            continue
        if c == "ú" or c == "ü":
            out.append("OO" if c == "ú" else "w")
            i += 1
            continue
        if c == "a":
            out.append("ah")
            i += 1
            continue
        if c == "e":
            out.append("eh")
            i += 1
            continue
        if c == "i" or c == "y":
            out.append("ee" if c == "i" or (c == "y" and not nxt) else "y")
            i += 1
            continue
        if c == "o":
            out.append("oh")
            i += 1
            continue
        if c == "u":
            out.append("oo")
            i += 1
            continue
        out.append(c)
        i += 1

    raw = "-".join(out)
    # compress
    raw = raw.replace("ah-ee", "eye").replace("eh-ee", "ay").replace("oh-ee", "oy")
    raw = raw.replace("ee-ah", "yah").replace("ee-eh", "yeh").replace("ee-oh", "yoh")
    raw = raw.replace("oo-ah", "wah").replace("oo-eh", "weh").replace("oo-ee", "wee")
    # stress last content syllable roughly: capitalize last chunk if multi
    parts = [p for p in raw.split("-") if p]
    if not parts:
        return word.upper()
    if len(parts) == 1:
        return parts[0]
    # default stress penultimate for vowel ending
    stress_i = len(parts) - 2 if word[-1] in "aeiounsáéíóú" else len(parts) - 1
    stress_i = max(0, min(stress_i, len(parts) - 1))
    parts[stress_i] = parts[stress_i].upper()
    return "-".join(parts)


def pluralize(noun: str) -> str:
    if noun.endswith("z"):
        return noun[:-1] + "ces"
    if noun[-1] in "aeiouáéíóú":
        return noun + "s"
    return noun + "es"


def adjective_forms(word: str) -> tuple[str, str, str, str]:
    """m.sg, f.sg, m.pl, f.pl"""
    if word.endswith("o"):
        stem = word[:-1]
        return word, stem + "a", stem + "os", stem + "as"
    if word.endswith("or"):
        return word, word + "a", word + "es", word + "as"
    if word.endswith("a"):
        # For words ending in 'a', treat as invariable
        pl = pluralize(word)
        return word, word, pl, pl
    if (
        word.endswith("e")
        or word.endswith("l")
        or word.endswith("z")
        or word.endswith("r")
        or word.endswith("n")
        or word.endswith("s")
    ):
        pl = pluralize(word)
        return word, word, pl, pl
    # For any other ending, treat as invariable
    pl = pluralize(word)
    return word, word, pl, pl


def load_freq() -> list[tuple[int, str, str]]:
    rows = []
    for line in FREQ.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        parts = line.split("\t")
        if len(parts) == 3:
            rank_s, surface, lemma = (p.strip() for p in parts)
            rows.append((int(rank_s), surface, lemma))
    return rows


def build() -> None:
    freq = load_freq()
    assert len(freq) == 1000, f"Expected 1000 rows, got {len(freq)}"

    wb = Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")
    verb_fill = PatternFill("solid", fgColor="E2EFDA")
    noun_fill = PatternFill("solid", fgColor="FFF2CC")
    adj_fill = PatternFill("solid", fgColor="FCE4D6")
    thin = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0D0"),
    )
    wrap = Alignment(wrap_text=True, vertical="center")

    # ----- Sheet: Read_Me -----
    ws0 = wb.active
    ws0.title = "Read_Me"
    notes = [
        ["Spanish Top 1000 Words — Study Workbook"],
        [""],
        [
            "Source frequency list: Wiktionary Spanish subtitle corpus (top 1000 word forms)."
        ],
        [
            "Present-tense conjugations generated for verb lemmas (regular + common irregulars)."
        ],
        [
            "Pronunciation: English-friendly guide (Latin America default), CAPS ≈ stress."
        ],
        [
            "Gender: m = masculine, f = feminine, mf = same form both, — = not applicable."
        ],
        [""],
        ["Sheets:"],
        [
            "1) Top_1000 — full frequency list (word form, lemma, type, gender, English, pronunciation,"
        ],
        [
            "   and verb persons: yo / tú / él-ella-usted / nosotros / vosotros / ellos-ustedes)."
        ],
        ["2) Study_Deck — ranks 1–200 only (less overwhelming daily practice)."],
        [
            "3) Verbs_Present — one row per person for every unique verb lemma (drill view)."
        ],
        ["4) Nouns_Adjectives — gender + singular/plural (and m/f adjective forms)."],
        [""],
        ["How to use:"],
        ["- Filter Type = verb on Top_1000 or Study_Deck to practice conjugations."],
        [
            "- For nouns: person columns are reused as article+singular / plural (see Notes column)."
        ],
        [
            "- Pair with spanish-cheatsheet.md for grammar explanations + click-to-hear words."
        ],
        ["- Rebuild this file: python scripts/build_top1000_excel.py"],
        [""],
        ["Persons (verbs, present indicative):"],
        ["  1st singular = yo"],
        ["  2nd singular = tú (informal)"],
        ["  3rd singular = él / ella / usted"],
        ["  1st plural = nosotros / nosotras"],
        ["  2nd plural = vosotros / vosotras (mainly Spain)"],
        ["  3rd plural = ellos / ellas / ustedes"],
        [""],
        ["Notes:"],
        [
            "- Some frequency rows are conjugated forms (e.g. estoy → lemma estar); conjugations follow the lemma."
        ],
        [
            "- Subtitle corpora include informal/vulgar slang; use judgment for classroom settings."
        ],
        ["- A few rare/irregular verbs may need a dictionary double-check."],
        ["- Pronunciation style matches spanish-cheatsheet.md."],
    ]
    for r in notes:
        ws0.append(r)
    ws0["A1"].font = Font(bold=True, size=14, color="2F5496")
    ws0.column_dimensions["A"].width = 110

    # ----- Sheet: Top_1000 -----
    ws = wb.create_sheet("Top_1000", 1)
    headers = [
        "Rank",
        "Word (form)",
        "Lemma",
        "Type",
        "Gender",
        "English",
        "Pronunciation (form)",
        "1st sg (yo)",
        "Pron. 1sg",
        "2nd sg (tú)",
        "Pron. 2sg",
        "3rd sg (él/ella/usted)",
        "Pron. 3sg",
        "1st pl (nosotros)",
        "Pron. 1pl",
        "2nd pl (vosotros)",
        "Pron. 2pl",
        "3rd pl (ellos/ustedes)",
        "Pron. 3pl",
        "Notes",
    ]
    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(1, col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(
            wrap_text=True, vertical="center", horizontal="center"
        )

    unique_verbs: dict[str, dict[str, str]] = {}
    nouns_adj: list[tuple] = []

    for rank, surface, lemma in freq:
        typ = guess_type(lemma, surface)
        gender = GENDER.get(lemma) or GENDER.get(surface) or "—"
        if typ == "verb":
            gender = "—"
        elif typ in (
            "preposition",
            "conjunction",
            "adverb",
            "pronoun",
            "article",
            "question word",
            "other",
        ):
            if gender == "—":
                gender = "—"
            if typ != "noun" and typ != "adjective" and lemma not in GENDER:
                gender = "—"

        meaning = english_meaning(lemma, surface, typ)

        pron_form = pronounce(surface)

        yo = tu = el = nos = vos = ellos = "—"
        p_yo = p_tu = p_el = p_nos = p_vos = p_ellos = "—"
        notes = ""

        conj = conjugate_regular(lemma) if typ == "verb" or lemma in IRREGULAR else None
        if conj is None and typ == "verb" and (lemma in IRREGULAR or is_verb(lemma)):
            conj = conjugate_regular(lemma)

        if conj and typ == "verb":
            gender = "—"
            yo, tu, el = conj["yo"], conj["tú"], conj["él"]
            nos, vos, ellos = conj["nosotros"], conj["vosotros"], conj["ellos"]
            p_yo, p_tu, p_el = pronounce(yo), pronounce(tu), pronounce(el)
            p_nos, p_vos, p_ellos = pronounce(nos), pronounce(vos), pronounce(ellos)
            unique_verbs[lemma] = conj
            if surface != lemma:
                notes = f"Surface form of lemma «{lemma}»"
            if lemma not in IRREGULAR and lemma not in E_IE | O_UE | E_I | U_UE:
                if re.search(r"(ar|er|ir)$", lemma):
                    notes = (
                        notes + "; " if notes else ""
                    ) + "Regular present (check stem-changers in dict if odd)"
        elif typ == "noun":
            g = (
                GENDER.get(lemma)
                or GENDER.get(surface)
                or ("f" if lemma.endswith("a") else "m")
            )
            gender = g if g != "mf" else "m/f"
            art = "la" if str(g).startswith("f") else "el"
            if lemma == "agua":
                art = "el"
                gender = "f (el agua)"
            pl = pluralize(lemma)
            yo = f"{art} {lemma}"
            tu = f"{'las' if art=='la' else 'los'} {pl}"
            p_yo = pronounce(lemma)
            p_tu = pronounce(pl)
            notes = "Noun: col I = article+sg; col K ≈ plural forms (not persons)"
            # reuse person columns as sg/pl for nouns for compact study
            el = pl
            p_el = p_tu
            nouns_adj.append(
                (lemma, "noun", gender, meaning, pronounce(lemma), f"{art} {lemma}", pl)
            )
        elif typ == "adjective":
            gender = GENDER.get(lemma) or "m/f"
            ms, fs, mp, fp = adjective_forms(lemma)
            yo, tu, el, nos = ms, fs, mp, fp
            p_yo, p_tu, p_el, p_nos = map(pronounce, (ms, fs, mp, fp))
            notes = (
                "Adjective: yo=m.sg, tú=f.sg, él=m.pl, nosotros=f.pl (not verb persons)"
            )
            nouns_adj.append(
                (lemma, "adjective", gender, meaning, pronounce(lemma), ms, fs)
            )
        else:
            notes = f"{typ}; person columns N/A"

        row = [
            rank,
            surface,
            lemma,
            typ,
            gender,
            meaning,
            pron_form,
            yo,
            p_yo,
            tu,
            p_tu,
            el,
            p_el,
            nos,
            p_nos,
            vos,
            p_vos,
            ellos,
            p_ellos,
            notes,
        ]
        ws.append(row)
        fill = None
        if typ == "verb":
            fill = verb_fill
        elif typ == "noun":
            fill = noun_fill
        elif typ == "adjective":
            fill = adj_fill
        if fill:
            for col in range(1, len(headers) + 1):
                ws.cell(ws.max_row, col).fill = fill
        for col in range(1, len(headers) + 1):
            ws.cell(ws.max_row, col).border = thin
            ws.cell(ws.max_row, col).alignment = wrap

    # widths
    widths = [
        6,
        14,
        14,
        12,
        12,
        28,
        18,
        14,
        14,
        14,
        14,
        18,
        14,
        16,
        14,
        16,
        14,
        18,
        14,
        36,
    ]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    ws.freeze_panes = "A2"

    # ----- Sheet: Study_Deck (ranks 1–200) -----
    ws_deck = wb.create_sheet("Study_Deck", 2)
    ws_deck.append(headers)
    for col, _ in enumerate(headers, 1):
        cell = ws_deck.cell(1, col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(
            wrap_text=True, vertical="center", horizontal="center"
        )
    for row in ws.iter_rows(min_row=2, max_row=201, values_only=True):
        ws_deck.append(list(row))
        r = ws_deck.max_row
        typ = row[3]
        fill = (
            verb_fill
            if typ == "verb"
            else (
                noun_fill if typ == "noun" else adj_fill if typ == "adjective" else None
            )
        )
        for col in range(1, len(headers) + 1):
            ws_deck.cell(r, col).border = thin
            ws_deck.cell(r, col).alignment = wrap
            if fill:
                ws_deck.cell(r, col).fill = fill
    for i, w in enumerate(widths, 1):
        ws_deck.column_dimensions[get_column_letter(i)].width = w
    ws_deck.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws_deck.max_row}"
    ws_deck.freeze_panes = "A2"

    # ----- Sheet: Verbs_Present -----
    ws2 = wb.create_sheet("Verbs_Present", 3)
    h2 = [
        "Lemma (infinitive)",
        "English",
        "Person",
        "Spanish pronoun",
        "Conjugated form",
        "Pronunciation",
        "Number",
    ]
    ws2.append(h2)
    for col, _ in enumerate(h2, 1):
        c = ws2.cell(1, col)
        c.font = header_font
        c.fill = header_fill

    person_meta = [
        ("1st", "yo", "yo", "singular"),
        ("2nd", "tú", "tú", "singular"),
        ("3rd", "él", "él / ella / usted", "singular"),
        ("1st", "nosotros", "nosotros / nosotras", "plural"),
        ("2nd", "vosotros", "vosotros / vosotras (Spain)", "plural"),
        ("3rd", "ellos", "ellos / ellas / ustedes", "plural"),
    ]

    # Ensure common verbs included even if not in unique from top1000
    for v in list(IRREGULAR.keys()):
        if v not in unique_verbs:
            unique_verbs[v] = IRREGULAR[v]

    # Add any verbs from VERB_MEANINGS that might not be in unique_verbs yet
    for v in list(VERB_MEANINGS.keys()):
        if v not in unique_verbs:
            conj = conjugate_regular(v)
            if conj:
                unique_verbs[v] = conj

    for lemma in sorted(
        unique_verbs.keys(), key=lambda x: (VERB_MEANINGS.get(x, x), x)
    ):
        conj = unique_verbs[lemma]
        eng = english_meaning(lemma, lemma, "verb")
        for person_label, key, pron_label, number in person_meta:
            form = conj[key]
            ws2.append(
                [
                    lemma,
                    eng,
                    person_label,
                    pron_label,
                    form,
                    pronounce(form),
                    number,
                ]
            )

    for i, w in enumerate([18, 28, 10, 28, 18, 18, 12], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.auto_filter.ref = f"A1:G{ws2.max_row}"
    ws2.freeze_panes = "A2"

    # ----- Sheet: Nouns_Adjectives -----
    ws3 = wb.create_sheet("Nouns_Adjectives", 4)
    h3 = [
        "Word",
        "Type",
        "Gender",
        "English",
        "Pronunciation",
        "Singular / m.sg",
        "Plural / f.sg or pl",
        "Extra (m.pl / f.pl)",
        "Notes",
    ]
    ws3.append(h3)
    for col, _ in enumerate(h3, 1):
        c = ws3.cell(1, col)
        c.font = header_font
        c.fill = header_fill

    seen_na = set()
    # from top 1000 nouns/adjectives
    for rank, surface, lemma in freq:
        typ = guess_type(lemma, surface)
        if typ not in ("noun", "adjective"):
            continue
        key = (lemma, typ)
        if key in seen_na:
            continue
        seen_na.add(key)
        gender = GENDER.get(lemma) or (
            "f" if lemma.endswith(("a", "ción", "sión", "dad", "tad")) else "m"
        )
        meaning = english_meaning(lemma, surface, typ)
        if typ == "noun":
            art = "la" if str(gender).startswith("f") and lemma != "mapa" else "el"
            if lemma == "agua":
                art = "el"
                gender = "f (uses el)"
            pl = pluralize(lemma)
            ws3.append(
                [
                    lemma,
                    typ,
                    gender,
                    meaning,
                    pronounce(lemma),
                    f"{art} {lemma}",
                    f"{'las' if art=='la' else 'los'} {pl}",
                    "",
                    "article + number",
                ]
            )
        else:
            ms, fs, mp, fp = adjective_forms(lemma)
            ws3.append(
                [
                    lemma,
                    typ,
                    "m/f",
                    meaning,
                    pronounce(lemma),
                    ms,
                    fs,
                    f"{mp} / {fp}",
                    "m.sg / f.sg / m.pl / f.pl",
                ]
            )

    # add extras from gender dict and noun_adj meanings
    for word in sorted(set(GENDER.keys()) | set(NOUN_ADJ_MEANINGS.keys())):
        if (word, "noun") in seen_na or (word, "adjective") in seen_na:
            continue
        g = GENDER.get(word)
        if g in ("m", "f") or g == "mf":
            # Determine if it's a noun or adjective
            typ = (
                "adjective"
                if word in NOUN_ADJ_MEANINGS and word not in GENDER
                else "noun"
            )
            if word in GENDER and word in NOUN_ADJ_MEANINGS:
                # If it's in both, check if it's an adjective by looking at common adjective endings
                if word.endswith(
                    ("o", "a", "e", "l", "z", "r", "n", "s")
                ) and word not in ["agua", "mano", "casa", "vida"]:
                    # It's likely an adjective
                    typ = "adjective"
            meaning = english_meaning(word, word, typ)
            if typ == "noun":
                art = "el" if g == "m" or word == "agua" else "la"
                if word == "agua":
                    art = "el"
                pl = pluralize(word)
                ws3.append(
                    [
                        word,
                        typ,
                        g,
                        meaning,
                        pronounce(word),
                        f"{art} {word}",
                        f"{'las' if art=='la' else 'los'} {pl}",
                        "",
                        "supplement",
                    ]
                )
            else:
                ms, fs, mp, fp = adjective_forms(word)
                ws3.append(
                    [
                        word,
                        typ,
                        "m/f",
                        meaning,
                        pronounce(word),
                        ms,
                        fs,
                        f"{mp} / {fp}",
                        "m.sg / f.sg / m.pl / f.pl",
                    ]
                )

    for i, w in enumerate([14, 12, 14, 24, 16, 18, 20, 18, 24], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.auto_filter.ref = f"A1:I{ws3.max_row}"
    ws3.freeze_panes = "A2"

    wb.save(OUT)
    print(f"Wrote {OUT}")
    print(f"Top_1000 rows: {ws.max_row - 1}")
    print(f"Study_Deck rows: {ws_deck.max_row - 1}")
    print(f"Unique verbs conjugated: {len(unique_verbs)}")
    print(f"Nouns_Adjectives rows: {ws3.max_row - 1}")


if __name__ == "__main__":
    build()
